import os

import ase
import numpy as np
import openpyxl
import rdkit
from ase.atom import Atom
from ase.atoms import Atoms
from ase.db import connect
from ase.io import read, write
from ase.neighborlist import build_neighbor_list
from ase.visualize import view
from openpyxl import load_workbook
from rdkit import Chem
from rdkit import Geometry
from rdkit.Chem import AllChem
from rdkit.Chem import rdMolAlign
from rdkit.Chem import rdmolfiles


def get_atoms_from_coords(file_path: str):
    flag = 0
    new_atoms = Atoms()
    with open(file_path) as f:
        for a_line in f.readlines():
            if a_line.startswith(' "coordinate'):
                flag = 1
                continue
            if a_line.startswith(' ],'):
                break
            if flag == 1:
                atoms_info = a_line.split('"')[1].split()
                an_atom = Atom(symbol=atoms_info[0],
                               position=(float(atoms_info[1]), float(atoms_info[2]), float(atoms_info[3])))
                new_atoms.append(an_atom)
    return new_atoms

# new_atoms = get_atoms_from_coords(r'ep-15_vum_sol.coordinate')
# a=1

def atom_2_mol(an_atoms: ase.atoms.Atoms):
    write(filename='temp.xyz', images=an_atoms)
    a_mol = rdmolfiles.MolFromXYZFile('temp.xyz')
    os.remove('temp.xyz')
    return a_mol


def set_mol_prop_from_row(row, mol):
    mol.SetProp('ep_id', f'{str(row.ep_id)}')
    mol.SetProp('smile', f'{row.smile}')
    mol.SetProp('binding_e', f'{str(row.binding_e)}')
    mol.SetProp('viscosity', f'{row.viscosity}')
    mol.SetProp('dielectric_constant', f'{row.dielectric_constant}')
    mol.SetProp('lumo', f'{row.lumo}')
    mol.SetProp('homo', f'{row.homo}')
    return mol


def mol_2_atom(mol: rdkit.Chem.rdchem.Mol):
    conf = mol.GetConformer()
    an_atoms = Atoms()
    for i in range(conf.GetNumAtoms()):
        position = conf.GetAtomPosition(i)
        atom = mol.GetAtoms()[i]
        a_symbol = atom.GetSymbol()
        an_new_atom = Atom(symbol=a_symbol, position=(position.x, position.y, position.z))
        an_atoms.append(an_new_atom)
    return an_atoms


def edit_ase_db(ase_db_path):
    with connect(ase_db_path) as db:
        meta = db.metadata
        if not "_distance_unit" in meta.keys():
            meta["_distance_unit"] = 'Ang'
        if not "_property_unit_dict" in meta.keys():
            meta["_property_unit_dict"] = {
                'viscosity': "Hartree",
                'dielectric_constant': "Hartree",  # relative
                'binding_e': "eV",
                'homo': "eV",
                'lumo': "eV",
            }
        if "atomrefs" not in meta.keys():
            meta["atomrefs"] = {}
        db.metadata = meta


def xlsx2db(db_path: str, xlsx_path: str, sheet_name: str):
    wb = load_workbook(filename=xlsx_path)
    ws = wb[sheet_name]

    real_count = 0
    fail_count = 0
    with connect(db_path) as db:
        for i, row in enumerate(ws.values):
            if i == 0:
                continue
            # if i>10:
            #     continue
            a_smile = row[2]
            try:
                a_mol = Chem.MolFromSmiles(a_smile)
                a_mol_with_H = Chem.AddHs(a_mol)
                AllChem.EmbedMolecule(a_mol_with_H, useRandomCoords=True, maxAttempts=1000000)
                AllChem.MMFFOptimizeMolecule(a_mol_with_H)

                an_atoms = mol_2_atom(mol=a_mol_with_H)

                an_id = row[1]
                an_id = int(an_id.split('-')[1])
                a_binding_e = float(row[3])
                a_dielectric_c = np.log10(float(row[4]))
                a_viscosity = np.log10(float(row[5]))
                a_lumo = float(row[6])
                a_homo = float(row[7])

                db.write(atoms=an_atoms, binding_e=a_binding_e, viscosity=a_viscosity,
                         dielectric_constant=a_dielectric_c, lumo=a_lumo, homo=a_homo,
                         smile=a_smile, ep_id=an_id)
                real_count = real_count + 1
                # print(f'real: {real_count}')
            except Exception as e:
                print(e)
                fail_count = fail_count + 1
                print(f'real: {real_count}')
                print(f'fail: {fail_count}')
                with open('fail_smile', 'a') as f_f:
                    f_f.write(a_smile)
                    f_f.write('\n')
    edit_ase_db(db_path)


def get_conf_from_db(smile: str, db_path: str=r'/root/submit_v2/ase_db/ood_test.db'):
    print('222222222222222222222')
    with connect(db_path) as db:
        q_atoms = db.get_atoms(smile=smile)
    q_mol = atom_2_mol(q_atoms)
    return q_mol.GetConformer()


def change_db_positions(old_db_path: str, new_db_path: str, coords_src_path: str):
    with connect(old_db_path) as old_db, connect(new_db_path) as new_db:
        for a_row in old_db.select():
            an_ep_id = a_row.ep_id
            coords_src_filename = f'ep-{an_ep_id}_vum_sol.coordinate'
            new_atoms = get_atoms_from_coords(os.path.join(coords_src_path, coords_src_filename))
            new_db.write(new_atoms, binding_e=a_row.binding_e, viscosity=a_row.viscosity,
                         dielectric_constant=a_row.dielectric_constant, lumo=a_row.lumo, homo=a_row.homo,
                         smile=a_row.smile, ep_id=an_ep_id)
    edit_ase_db(new_db_path)

def change_db_property(old_db_path: str, new_db_path: str, enlarge: bool=True):
    with connect(old_db_path) as old_db, connect(new_db_path) as new_db:
        for a_row in old_db.select():
            an_ep_id = a_row.ep_id
            new_atoms = old_db.get_atoms(ep_id=an_ep_id)
            a_dc = a_row.dielectric_constant
            
            if enlarge:
                new_dc = pow(10, a_dc)
            else:
                new_dc = np.log10(a_dc)

            new_db.write(new_atoms, binding_e=a_row.binding_e, viscosity=a_row.viscosity,
                         dielectric_constant=new_dc, lumo=a_row.lumo, homo=a_row.homo,
                         smile=a_row.smile, ep_id=an_ep_id)
    edit_ase_db(new_db_path)


def get_rmsd_between_2_db(db_1_path: str, db_2_path: str):
    rmsd_list = []
    with connect(db_1_path) as db_1, connect(db_2_path) as db_2:
        for a_row in db_1.select():
            an_ep_id = a_row.ep_id
            an_atoms_1 = db_1.get_atoms(ep_id=an_ep_id)
            an_atoms_2 = db_2.get_atoms(ep_id=an_ep_id)
            mol_1 = atom_2_mol(an_atoms_1)
            mol_2 = atom_2_mol(an_atoms_2)
            a_rmsd = rdMolAlign.AlignMol(mol_1, mol_2)
            rmsd_list.append(a_rmsd)
    return rmsd_list


def xlsx2xyz(dump_path: str, xlsx_path: str, sheet_name: str):
    wb = load_workbook(filename=xlsx_path)
    ws = wb[sheet_name]

    cwd_ = os.getcwd()
    fail_smile_path = os.path.abspath('fail_smile')
    real_smile_path = os.path.abspath('real_smile')

    os.makedirs(dump_path, exist_ok=True)
    os.chdir(dump_path)

    real_count = 0
    fail_count = 0
    for i, row in enumerate(ws.values):
        if i == 0:
            continue

        a_smile = row[0]
        if a_smile == None:
            print(i)
            continue
################################################################################################
        try:
            a_mol = Chem.MolFromSmiles(a_smile)
            a_mol_with_H = Chem.AddHs(a_mol)
            AllChem.EmbedMolecule(a_mol_with_H, useRandomCoords=True, maxAttempts=1000000)
            AllChem.MMFFOptimizeMolecule(a_mol_with_H)

            an_atoms = mol_2_atom(mol=a_mol_with_H)
            new_name = a_smile + f'_{str(len(an_atoms))}' + '.xyz'
            if new_name in os.listdir(dump_path):
                print(a_smile)
                continue

            write(filename=new_name, format='xyz', images=an_atoms)
            with open(real_smile_path, 'a') as f_a:
                f_a.write(a_smile)
                f_a.write('\n')
        except Exception as e:
            print(e)
            fail_count = fail_count + 1
            print(f'real: {real_count}')
            print(f'fail: {fail_count}')

            print(row)
            print(a_smile)

            with open(fail_smile_path, 'a') as f_f:
                f_f.write(a_smile)
                f_f.write('\n')
################################################################################################
    os.chdir(cwd_)


def get_new_db(old_db_path: str, new_db_path: str):
    with connect(old_db_path) as old_db, connect(new_db_path) as new_db:
        for a_row in old_db.select():
            an_ep_id = a_row.ep_id
            new_atoms = old_db.get_atoms(ep_id=an_ep_id)
            data = {
                'binding_e': [a_row.binding_e],
                'viscosity': [a_row.viscosity],
                'dielectric_constant': [a_row.dielectric_constant],
                'lumo': [a_row.lumo],
                'homo': [a_row.homo],
            }
            new_db.write(new_atoms, data=data, ep_id=a_row.ep_id, smile=a_row.smile)
    edit_ase_db(new_db_path)


os.makedirs('./data3')
for a_db in os.listdir('./data'):
    tgt = os.path.join('./data3', a_db)
    src = os.path.join('./data', a_db)
    get_new_db(old_db_path=src, new_db_path=tgt)



